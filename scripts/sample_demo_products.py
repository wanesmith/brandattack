"""
Sample ~50 SKUs from the seed Adidas xlsx and emit:
  - src/data/products.json — product + variant records for the demo
  - scripts/_image_filelist.txt — list of "<ArticleNo>-<N>.jpg" filenames to extract from the 7z

Run from project root: py scripts/sample_demo_products.py
"""

import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

XLSX = Path(r"C:\Users\wane\Downloads\Adidas (21,082 Units)-VF - For Wane.xlsx")
ROOT = Path(__file__).resolve().parent.parent
OUT_JSON = ROOT / "src" / "data" / "products.json"
OUT_IMG_LIST = ROOT / "scripts" / "_image_filelist.txt"

SAMPLE_TARGET = 60  # we ask for 60 to allow drop-outs (missing images, no variants, etc.)

# --- Normalisation helpers -------------------------------------------------

GENDER_MAP = {
    "FEMALE": "WOMEN", "WOMEN": "WOMEN",
    "MALE": "MEN", "MEN": "MEN",
    "UNISEX": "UNISEX",
    "KIDS": "KIDS",
}

SIZE_LABEL_RE = re.compile(r"^(\d+)(-)?(K)?$")


def normalise_gender(g: str | None) -> str:
    if not g:
        return "UNISEX"
    return GENDER_MAP.get(g.upper().strip(), "UNISEX")


def size_label(raw: str) -> str:
    """Turn raw size codes into something human-readable.
    Source sizes are mixed conventions (US footwear, EU kids in cm, alpha apparel)
    so don't claim a region — just decode the trailing '-' as a .5 step and the
    'K' suffix as kids.
        '5'  -> '5'
        '5-' -> '5.5'
        '10K'-> '10 (kids)'
        '5-K'-> '5.5 (kids)'
        'M' / 'XL' / '104' stay as-is.
    """
    raw = str(raw).strip()
    m = SIZE_LABEL_RE.match(raw)
    if m:
        num, half, kids = m.group(1), m.group(2), m.group(3)
        n = f"{num}.5" if half else num
        return f"{n} (kids)" if kids else n
    return raw


def title_case(s: str | None) -> str:
    if not s:
        return ""
    # The data is mostly ALL CAPS — title-case it but leave 1–2 letter tokens (like "H", "W") alone.
    return " ".join(w.capitalize() if len(w) > 2 else w.upper() for w in s.split())


def slugify(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


# --- Pricing ----------------------------------------------------------------

DEFAULT_MARKUP_OVER_COST = 2.0  # retail = 2x wholesale, configurable later


def compute_retail_usd(wholesale: float, rrp: float) -> float:
    """Demo pricing: 2x cost, capped at 70% of RRP so 'X% off' looks credible."""
    candidate = wholesale * DEFAULT_MARKUP_OVER_COST
    return round(min(candidate, rrp * 0.7), 2)


# --- Main -------------------------------------------------------------------

def main():
    print(f"Reading {XLSX} ...")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    asset = wb["Asset Report"]
    vertical = wb["VerticaL Article Size Qty"]
    pics = wb["Pictures List"]

    # Build images-by-article map from sheet 3
    images_by_article: dict[str, list[str]] = defaultdict(list)
    for row in pics.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 2 or not row[0] or not row[1]:
            continue
        fn, art = str(row[0]).strip(), str(row[1]).strip()
        if fn.lower().endswith(".jpg"):
            images_by_article[art].append(fn)
    for art in images_by_article:
        images_by_article[art].sort()  # BA8044-1.jpg before BA8044-2.jpg

    # Build variants from sheet 2 (vertical)
    variants_by_article: dict[str, list[tuple[str, int]]] = defaultdict(list)
    for i, row in enumerate(vertical.iter_rows(min_row=2, values_only=True)):
        if not row or not row[0]:
            continue
        art, attr, qty = row[0], row[1], row[2]
        if not attr or qty is None:
            continue
        try:
            qty_i = int(qty)
        except (TypeError, ValueError):
            continue
        if qty_i <= 0:
            continue
        variants_by_article[str(art).strip()].append((str(attr).strip(), qty_i))

    # Walk asset report, header is row 2
    header = [c.value for c in next(asset.iter_rows(min_row=2, max_row=2))]
    idx = {h: i for i, h in enumerate(header) if h}

    needed = ["ArticleNo", "Item Description", "Season", "Division", "Gender",
              "Sports Code", "Product Group", "Product Type",
              "Total Quantity", "Discount", "Price", "RRP Price (USD)"]
    for col in needed:
        if col not in idx:
            raise RuntimeError(f"Missing column {col!r} in xlsx")

    # Collect candidate products with the metadata we need
    candidates = []
    for row in asset.iter_rows(min_row=3, values_only=True):
        art = row[idx["ArticleNo"]]
        if not art:
            continue
        art = str(art).strip()
        if not images_by_article.get(art):
            continue  # skip SKUs with no photos
        if not variants_by_article.get(art):
            continue  # skip SKUs with no in-stock sizes
        candidates.append({
            "articleNo": art,
            "description": (row[idx["Item Description"]] or "").strip(),
            "season": row[idx["Season"]] or "",
            "division": (row[idx["Division"]] or "").upper(),
            "gender_raw": row[idx["Gender"]] or "",
            "sportsCode": (row[idx["Sports Code"]] or "").upper(),
            "productGroup": (row[idx["Product Group"]] or "").upper(),
            "productType": (row[idx["Product Type"]] or "").upper(),
            "totalQty": int(row[idx["Total Quantity"]] or 0),
            "wholesaleUsd": float(row[idx["Price"]] or 0),
            "rrpUsd": float(row[idx["RRP Price (USD)"]] or 0),
        })

    print(f"  {len(candidates)} viable SKUs (have images + variants).")

    # Stratified sample: spread across (division, normalised gender)
    buckets: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for c in candidates:
        key = (c["division"], normalise_gender(c["gender_raw"]))
        buckets[key].append(c)

    # Round-robin from buckets so the demo looks varied
    per_bucket = max(1, SAMPLE_TARGET // len(buckets))
    sample: list[dict] = []
    # sort buckets so result is deterministic
    for key in sorted(buckets):
        sample.extend(buckets[key][:per_bucket])
    # if we're short, pull more from the largest buckets
    if len(sample) < SAMPLE_TARGET:
        for key in sorted(buckets, key=lambda k: -len(buckets[k])):
            extra_needed = SAMPLE_TARGET - len(sample)
            if extra_needed <= 0:
                break
            sample.extend(buckets[key][per_bucket:per_bucket + extra_needed])
    sample = sample[:SAMPLE_TARGET]

    print(f"  Sampled {len(sample)} across {len(buckets)} (division, gender) buckets.")

    # Build output records
    products = []
    needed_images: set[str] = set()
    for c in sample:
        art = c["articleNo"]
        gender = normalise_gender(c["gender_raw"])
        title = title_case(c["description"]) or art
        slug = f"{art.lower()}-{slugify(title)}"
        rrp = c["rrpUsd"]
        price = compute_retail_usd(c["wholesaleUsd"], rrp)
        # Limit demo to first 4 images per product (more than enough)
        imgs = images_by_article[art][:4]
        for fn in imgs:
            needed_images.add(fn)
        variants = [
            {
                "size": raw,
                "sizeLabel": size_label(raw),
                "stock": qty,
                "sku": f"{art}-{raw.replace('/', '_')}",
            }
            for raw, qty in variants_by_article[art]
        ]
        products.append({
            "id": slug,
            "articleNo": art,
            "title": title,
            "description": c["description"],
            "brand": "Adidas",
            "division": c["division"],
            "gender": gender,
            "sportsCode": c["sportsCode"],
            "productGroup": c["productGroup"],
            "productType": c["productType"],
            "season": c["season"],
            "rrpUsd": rrp,
            "priceUsd": price,
            "totalStock": sum(v["stock"] for v in variants),
            "images": [f"/products/{Path(fn).stem}.webp" for fn in imgs],
            "variants": variants,
        })

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(products, indent=2), encoding="utf-8")
    print(f"  Wrote {OUT_JSON} ({len(products)} products).")

    OUT_IMG_LIST.write_text("\n".join(sorted(needed_images)), encoding="utf-8")
    print(f"  Wrote {OUT_IMG_LIST} ({len(needed_images)} images needed).")


if __name__ == "__main__":
    main()
